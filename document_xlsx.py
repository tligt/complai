"""
document_xlsx.py — S26

Builds an XLSX from a rendered document's structured blocks.

WHY XLSX IS THE PRIMARY FORMAT FOR A REGISTER (D-29)
-----------------------------------------------------
A RoPA is a table, not an essay. It gets filtered, sorted and handed to an
authority — the CNIL publishes its own model register as a spreadsheet, and
that is the artefact a French client's auditor will recognise. A twelve-column
landscape Word table is technically a record and practically unusable.

DOCX stays available and is built from the same render. Same content, same
template version, two representations, ONE documents row (D-29). Never two
rows for one record, or S27 tracks two adoption states for one document.

WHAT THIS CONSUMES
------------------
RenderResult.blocks — the structured form collected during rendering. NOT the
markdown body: parsing our own table back into cells would mean unescaping the
pipes _cell() escaped, and the escaping is a markdown concern that must never
reach a spreadsheet cell.

The prose around the table is not discarded. It becomes the "About" sheet, so
the scope statement and the as-at date travel with the file rather than living
only in the DOCX.
"""

from __future__ import annotations

import io
import re
from typing import Any, Sequence

from template_renderer import Block


# Approximate character widths per column, by header position. Openpyxl has no
# autofit, and a register whose columns are all 8 characters wide is unreadable
# in exactly the situation it exists for.
_MIN_WIDTH = 14
_MAX_WIDTH = 60


def _sheet_title(name: str) -> str:
    """Excel sheet names: 31 chars, and none of : \\ / ? * [ ]"""
    cleaned = re.sub(r"[:\\/?*\[\]]", " ", name).strip() or "Sheet"
    return cleaned[:31]


def _strip_markdown(text: str) -> str:
    """Prose for the About sheet, with markdown emphasis removed.

    Headings keep their text and lose the hashes; bold and italic markers go.
    Deliberately crude — this is a readable summary, not a second renderer.
    """
    out: list[str] = []
    for raw in text.splitlines():
        line = raw.rstrip()
        if line.lstrip().startswith("|"):
            continue  # the table has its own sheet
        line = re.sub(r"^#{1,6}\s*", "", line)
        line = line.replace("**", "").replace("__", "")
        line = re.sub(r"(?<!\w)[*_](?=\S)(.+?)(?<=\S)[*_](?!\w)", r"\1", line)
        out.append(line)
    # Collapse the blank runs the stripped table leaves behind.
    collapsed: list[str] = []
    for line in out:
        if not line.strip() and collapsed and not collapsed[-1].strip():
            continue
        collapsed.append(line)
    return "\n".join(collapsed).strip()


def build_xlsx(
    blocks: Sequence[Block],
    *,
    title: str,
    prose: str | None = None,
    metadata: Sequence[tuple[str, Any]] = (),
) -> bytes:
    """One sheet per block, plus an About sheet carrying the prose.

    Raises ImportError with a usable message if openpyxl is absent, rather than
    failing at import time and taking the whole page down — the DOCX path must
    keep working on an environment where the dependency has not landed yet.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required to export a register as a spreadsheet. "
            "Add `openpyxl` to requirements.txt."
        ) from exc

    wb = Workbook()

    # --- About sheet -------------------------------------------------------
    about = wb.active
    about.title = _sheet_title("About")
    about["A1"] = title
    about["A1"].font = Font(bold=True, size=14)

    row = 3
    for label, value in metadata:
        about.cell(row=row, column=1, value=label).font = Font(bold=True)
        about.cell(row=row, column=2, value=value)
        row += 1

    if prose:
        row += 1
        for line in _strip_markdown(prose).splitlines():
            cell = about.cell(row=row, column=1, value=line or None)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    about.column_dimensions["A"].width = 34
    about.column_dimensions["B"].width = 70

    # --- One sheet per block ----------------------------------------------
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="14C7D5")
    wrap = Alignment(wrap_text=True, vertical="top")

    for block in blocks:
        ws = wb.create_sheet(_sheet_title(block.name.replace("_", " ").title()))

        ws.append(list(block.headers))
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        for r in block.rows:
            # None becomes an empty cell, not the string "None". An em dash is
            # a markdown affordance; a spreadsheet filter should see a blank.
            ws.append([("" if v is None else v) for v in r])

        for idx, header in enumerate(block.headers, start=1):
            longest = max(
                [len(str(header))]
                + [len(str(r[idx - 1])) for r in block.rows
                   if idx - 1 < len(r) and r[idx - 1] is not None]
                or [0]
            )
            ws.column_dimensions[get_column_letter(idx)].width = max(
                _MIN_WIDTH, min(_MAX_WIDTH, longest + 2)
            )

        for line in ws.iter_rows(min_row=2):
            for cell in line:
                cell.alignment = wrap

        # Filters and a frozen header: the two things that make a register
        # usable when someone is looking for one row among sixty.
        if block.rows:
            ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        if block.caption:
            ws.append([])
            ws.append([_strip_markdown(block.caption)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
